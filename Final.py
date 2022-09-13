##python 3.8

import seaborn as sns
import pandas as pd
import xlrd
import matplotlib.pyplot as plt
import wikipedia
from openpyxl import Workbook
import math as maths
from bokeh.plotting import show
import urllib.request, urllib.error, urllib.parse
from bs4 import BeautifulSoup
import requests
import holoviews as hv
from holoviews import opts, dim
sns.set_theme()



#problem 1

webpages = [[]for i in range(10)]
keywords = ["targeted threat", "Advanced Persistent Threat", "phishing", "DoS attack", "malware", "computer virus", "spyware", "malicious bot", "ransomware", "encryption"]
#keywords = ["Advanced Persistent Threat"]
for y in range(0, 10):
    n = 1
    while len(webpages[y]) <= 100:
        
        search = requests.get('https://www.bbc.co.uk/search?q='+ keywords[y] + '&page=' + str(n))
        searchSoup = BeautifulSoup(search.content, 'html.parser')
        data = searchSoup.find_all('div', class_ = "ssrcss-1bn8j6y-PromoContent e1f5wbog0")
        if len(data) == 0:
            break
        for x in data:
            data2 = x.find_all('a', class_ = "ssrcss-10ivm7i-PromoLink e1f5wbog5")
            for a in data2:
                #print(a['href'])
                page = a['href']

                if not "https://www.bbc.co.uk/programmes" in page and not "www.bbc.co.uk/bitesize" in page:
                    webpages[y].append(page)

        n = n + 1

print(webpages[1])

pagesContents = [[]for i in range(10)]
for y in range(0, 10):
    for x in range(0, len(webpages[y])):
        print(x)
        url = webpages[y][x]
        try:
            response = urllib.request.urlopen(url)
            webContent = response.read()
        except:
            print("Not found")

        pagesContents[y].append(webContent)
        #f = open(str(x) + '.html', 'wb')
        #f.write(webContent)
        #f.close


#problem 2
for y in range(0, 10):
    for n in range(0, len(webpages[y])):
        if n >= 100:
            break
        soup = BeautifulSoup(pagesContents[y][n], 'html.parser')
        #print(soup.get_text())
        f = open(keywords[y] + str(n) + '.txt', 'w')
        #f.write(soup.get_text())
        #f.write(webpages[y][n])
        if webpages[y][n].endswith(".stm"):
            #f.write("Old Page")
            data = soup.find_all('p')
            for x in data:
                #print(x.get_text())
                try:
                    f.write(x.get_text())
                except:
                    print("text not working")
        else:
            data = soup.find_all('div', class_ = "ssrcss-uf6wea-RichTextComponentWrapper e1xue1i83")

            for x in data:
                #print(x.get_text())
                try:
                    f.write(x.get_text())
                except:
                    print("text not working")
        f.close



#problem 3
menu = 0
#keywords = ["targeted threat", "Advanced Persistent Threat", "phishing", "DoS attack", "malware", "computer virus", "spyware", "malicious bot", "ransomware", "encryption"]
BBCDistances = [[0] * 10 for a in range(10)]
wikiDistances = [[0] * 10 for a in range(10)]
distances = [[0] * 10 for a in range(10)]
BBCOccurances = [[0] * 10 for a in range(10)]
WikiOccurances = [[0] * 10 for a in range(10)]

def vectorMagnitude(vector):
    magnitude = 0
    for value in vector:
        magnitude = magnitude + (value ** 2)
    magnitude = maths.sqrt(magnitude)
    return magnitude

    
def vectorMultiplication(vector1, vector2):
    output = 0
    for y in range(1, 10):
        output = output + (vector1[y] * vector2[y]) 
    return output    


def bbcArticles():

    #calculate number of occurances
    for x in range(0, 10):
        
        for i in range(0, 100):
            try:
                f = open(keywords[x] + str(i) + '.txt', "r")
                content = f.read()
                for y in range(0, 10):
                    BBCOccurances[x][y] = BBCOccurances[x][y] + content.count(keywords[y])
            except:
                pass
    #print(BBCOccurances)
    
    for x in range(0, 10):
        for y in range(0, 10):
                BBCDistances[x][y] = (vectorMultiplication(BBCOccurances[x], BBCOccurances[y]) / (vectorMagnitude(BBCOccurances[x]) * vectorMagnitude(BBCOccurances[y])))


def wikipediaArticles():
    wikiArts = []
    
    for x in range(0, 10):
        searchTerms = []
        #print(keywords[x])
        searchTerms = wikipedia.search(keywords[x], results=5)
        for i in range(0, 5):
            t = wikipedia.page(searchTerms[i], auto_suggest=False).content
            for y in range(0, 10):
                WikiOccurances[x][y] = WikiOccurances[x][y] + t.count(keywords[y])
    for x in range(0, 10):
        for y in range(0, 10):
                wikiDistances[x][y] = (vectorMultiplication(WikiOccurances[x], WikiOccurances[y]) / (vectorMagnitude(WikiOccurances[x]) * vectorMagnitude(WikiOccurances[y])))
        

wikipediaArticles()
bbcArticles()
#distances = BBCDistances
#distances = wikiDistances



book = Workbook()
sheet = book.active


for x in range(2, 12):
    sheet.cell(row=1, column=x).value = keywords[x - 2]
    sheet.cell(row=x, column=1).value = keywords[x - 2]
    for y in range(2, 12):
        distances[x - 2][y - 2] = (BBCDistances[x - 2][y - 2] +  wikiDistances[x - 2][y - 2]) / 2
        sheet.cell(row=x, column=y).value = distances[x - 2][y - 2]
#print(distances)
book.save("distance.xlsx")


#problem 4
#heatmap
semDistanceData = pd.read_excel('distance.xlsx', engine='openpyxl', index_col=0)
sns.heatmap(semDistanceData, annot=True).set_title('Semantic Distance from BBC and Wikipedia Articles')#, fmt = ".2f")#, fmt="d")
plt.show()

#chord graph
hv.extension('bokeh', 'matplotlib')
hv.output(size=200)

n = 45
m = 3
chordValues = [[0] * m for i in range(n)]
i = 0

for x in range(0, 9):
    for y in range(x + 1, 10):
        chordValues[i][0] = x
        chordValues[i][1] = y
        chordValues[i][2] = semDistanceData.iat[x,y]
        i = i + 1
        
df = pd.DataFrame(chordValues)
df.columns = ['source', 'target', 'value']

hv.Chord(df)
nodes = []
for x in range(0, 10):
    nodes.append([keywords[x]])

nodesDf = pd.DataFrame(data = nodes)
nodesDf.columns =['keyword']

nodesDset = hv.Dataset(nodesDf, 'index')
#print(nodesDset)

chord = hv.Chord((df, nodesDset))

chord.opts(
    opts.Chord(cmap='Category20', edge_cmap='Category20', edge_color=dim('source').str(), 
               labels='keyword',title="Semantic Distance from BBC and Wikipedia Articles", node_color=dim('index').str()))

show(hv.render(chord))        


